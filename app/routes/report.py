
from flask import Blueprint, render_template, jsonify, request, redirect, Response, flash, url_for, current_app, send_file
from flask_login import login_required
from sqlalchemy import cast, Integer, func
from sqlalchemy.orm import joinedload
from app import db
from app.routes.decorators import require_module
from datetime import datetime, timedelta
import os, json, requests, shutil
from xlrd import open_workbook
from xlutils.copy import copy
from openpyxl import load_workbook, Workbook
from sqlalchemy import func, or_, and_
from collections import OrderedDict
from app.models import CTMS1000, CTMS4100, Cases, CTMS4000
from dateutil.relativedelta import relativedelta
from datetime import timedelta
from openpyxl.utils import get_column_letter
from sqlalchemy import cast, String
from collections import Counter
from openpyxl.styles import Font, Alignment, Border, Side
from sqlalchemy import case






reports_bp = Blueprint('reports', __name__, url_prefix='/report')



civil_list = {
    "10001": "Small Claims",
    "10002": "Ejectment",
    "10003": "Other Cases under Summary Procedure",
    "10004": "Cases under Regular Procedure",
    "10005": "Election",
    "10006": "Other Civil Cases"
}


def classify_civil_case(title):
    if not title:
        return "10006"  # Other Civil Cases

    text = title.upper().strip()

    # Ejectment
    if any(x in text for x in [
        "UNLAWFUL DETAINER",
        "FORCIBLE ENTRY"
    ]):
        return "10002"

    # Election
    if any(x in text for x in [
        "ELECTION",
        "ELECTION PROTEST",
        "QUO WARRANTO"
    ]):
        return "10005"

    # Small Claims
    if "SMALL CLAIM" in text:
        return "10001"

    # Other Summary Procedure
    if "SUMMARY PROCEDURE" in text:
        return "10003"

    # Default
    return "10004"  # Cases under Regular Procedure

def get_civil_case_counts(query):
    """
    query must return Cases.nature
    """

    counter = Counter()

    for (nature,) in query.all():
        code = classify_civil_case(nature)
        counter[code] += 1

    return [
        (code, counter.get(code, 0))
        for code in civil_list.keys()
    ]


@reports_bp.route('/')
@login_required
@require_module(16)
def index():

    storage_path = os.path.join(
        current_app.root_path,
        'static',
        'storage'
    )

    os.makedirs(storage_path, exist_ok=True)

    files = []

    for filename in os.listdir(storage_path):

        full_path = os.path.join(storage_path, filename)

        if os.path.isfile(full_path):

            files.append({
                "name": filename,
                "size": round(os.path.getsize(full_path) / 1024, 2),
                "date": datetime.fromtimestamp(
                    os.path.getmtime(full_path)
                ).strftime("%Y-%m-%d %I:%M %p")
            })

    # newest first
    files.sort(key=lambda x: x["date"], reverse=True)

    return render_template(
        "reports/index.html",
        files=files
    )




@reports_bp.route('/delete/<filename>')
@login_required
@require_module(16)
def delete_file(filename):

    storage_path = os.path.join(
        current_app.root_path,
        'static',
        'storage'
    )

    file_path = os.path.join(storage_path, filename)

    if os.path.exists(file_path):
        os.remove(file_path)
        flash(f"{filename} deleted successfully.", "success")
    else:
        flash("File not found.", "danger")

    return redirect(url_for('reports.index'))


@reports_bp.route('/mrc')
@login_required
@require_module(16)
def mrc():

    # =========================
    # PATHS
    # =========================
    base_file = os.path.join(
        current_app.root_path,
        "static",
        "base_docs",
        "MTC BUENAVISTA, AGUSAN DEL NORTE.xlsx"
    )

    storage_dir = os.path.join(
        current_app.root_path,
        "static",
        "storage"
    )

    os.makedirs(storage_dir, exist_ok=True)

    # =========================
    # DATE FORMAT
    # =========================
    month_param = request.args.get("month")

    if month_param:
        now = datetime.strptime(month_param, "%Y-%m")
    else:
        now = datetime.now()

    month_year_text = now.strftime("%B %Y")

 
    month_start = datetime(now.year, now.month, 1)
   

    next_month_start = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)


    # =========================
    # FILE NAME
    # =========================
    base_filename = f"{month_year_text} MTC BUENAVISTA, AGUSAN DEL NORTE.xlsx"

    name, ext = os.path.splitext(base_filename)

    counter = 1
    filename = base_filename

    while os.path.exists(os.path.join(storage_dir, filename)):
        filename = f"{name} ({counter}){ext}"
        counter += 1

    output_path = os.path.join(storage_dir, filename)

    shutil.copy2(base_file, output_path)
    

    # =========================
    # ROW CONFIG
    # =========================
    rows = [14, 16, 17, 18, 20, 21, 22, 23]

    nature_list = ["00027","00028","00029","00030","00031","00032","00033","00034","00035","00036"]
    civil_list = ["10001","10002","10003","10004","10005","10006"]
    start_col = 5  

    # =========================
    # PAGE 1 Counting Result
    # =========================

    p1_row16_result = (
        db.session.query(
            CTMS1000.NATURECODE,
            func.count(CTMS1000.NATURECODE)
        )
        .filter(
            CTMS1000.DTFILED.like(f"{now.strftime('%Y-%m')}%")
        )
        .group_by(CTMS1000.NATURECODE)
        .all()
    )

    p1_row17_result = (
        db.session.query(
            CTMS1000.NATURECODE,
            func.count(CTMS1000.NATURECODE)
        )
        .join(CTMS4100, CTMS4100.CASEID == CTMS1000.CASEID)
        .filter(
            CTMS4100.DTREVIVED.like(f"{now.strftime('%Y-%m')}%")
        )
        .group_by(CTMS1000.NATURECODE)
        .all()
    )

    p1_row20_result = (
        db.session.query(
            CTMS1000.NATURECODE,
            func.count(CTMS1000.NATURECODE)
        )
        .join(CTMS4100, CTMS4100.CASEID == CTMS1000.CASEID)
        .filter(
            CTMS4100.DECIDECODE == "90002",
            CTMS4100.DTPROMUL.like(f"{now.strftime('%Y-%m')}%"),
            CTMS4100.DECIDETYPE == "ON_MERITS"
        )
        .group_by(CTMS1000.NATURECODE)
        .all()
    )

    p1_row21_result = (
        db.session.query(
            CTMS1000.NATURECODE,
            func.count(CTMS1000.NATURECODE)
        )
        .join(CTMS4100, CTMS4100.CASEID == CTMS1000.CASEID)
        .filter(
            CTMS4100.DECIDECODE != "90002",
            CTMS4100.DTPROMUL.like(f"{now.strftime('%Y-%m')}%"),
            CTMS4100.DECIDETYPE != "ON_MERITS"
        )
        .group_by(CTMS1000.NATURECODE)
        .all()
    )



    p1_row22_result = (
        db.session.query(
            CTMS1000.NATURECODE,
            func.count(CTMS1000.NATURECODE)
        )
        .join(CTMS4100, CTMS4100.CASEID == CTMS1000.CASEID)
        .filter(
            CTMS4100.DTARCHIVED.like(f"{now.strftime('%Y-%m')}%")
        )
        .group_by(CTMS1000.NATURECODE)
        .all()
    )

    p1_row23_result = (
        db.session.query(
            CTMS1000.NATURECODE,
            func.count(CTMS1000.NATURECODE)
        )
        .join(CTMS4100, CTMS4100.CASEID == CTMS1000.CASEID)
        .filter(
            CTMS4100.DTREFERRED.like(f"{now.strftime('%Y-%m')}%")
        )
        .group_by(CTMS1000.NATURECODE)
        .all()
    )

    p1_row14_result = (
        db.session.query(
            CTMS1000.NATURECODE,
            func.count(CTMS1000.NATURECODE)
        )
        .join(CTMS4100, CTMS4100.CASEID == CTMS1000.CASEID)
        .group_by(CTMS1000.NATURECODE)
        .all()
    )


    # =========================
    # PAGE 2 Counting Result
    # =========================

    query = (
        db.session.query(Cases.nature)
        .filter(
            Cases.date_filed.like(f"{now.strftime('%Y-%m')}%")
        )
    )

    p2_row16_result = get_civil_case_counts(query)

    query = (
        db.session.query(Cases.nature)
        .filter(
            func.json_extract(Cases.action, '$.revived_date').like(f"{now.strftime('%Y-%m')}%"),
        )
    )
    p2_row17_result = get_civil_case_counts(query)



    query = (
        db.session.query(Cases.nature)
        .filter(
            func.json_extract(Cases.action, '$.decision_date').like(f"{now.strftime('%Y-%m')}%"),

            func.json_extract(Cases.action, '$.decision_type') == "ON_MERIT",

        )
    )
    p2_row20_result = get_civil_case_counts(query)

    query = (
        db.session.query(Cases.nature)
        .filter(
            func.json_extract(Cases.action, '$.decision_date').like(f"{now.strftime('%Y-%m')}%"),

            func.json_extract(Cases.action, '$.decision_type') != "ON_MERIT",

        )
    )
    p2_row21_result = get_civil_case_counts(query)

    query = (
        db.session.query(Cases.nature)
        .filter(
            func.json_extract(Cases.action, '$.archived_date').like(f"{now.strftime('%Y-%m')}%"),

        )
    )
    p2_row22_result = get_civil_case_counts(query)
 

    query = (
        db.session.query(Cases.nature)
        .filter(
            func.json_extract(Cases.action, '$.referred_date').like(f"{now.strftime('%Y-%m')}%"),

        )
    )
    p2_row23_result = get_civil_case_counts(query)

    # =========================
    # COUNT DICTS
    # =========================

    p1_counts = {
        14: dict(p1_row14_result),
        16: dict(p1_row16_result),
        17: dict(p1_row17_result),
        18: {},
        20: dict(p1_row20_result),
        21: dict(p1_row21_result),
        22: dict(p1_row22_result),
        23: dict(p1_row23_result),
    }


    p2_counts = {
        # 14: dict(p1_row14_result),
        16: dict(p2_row16_result),
        17: dict(p2_row17_result),
        18: {},
        20: dict(p2_row20_result),
        21: dict(p2_row21_result),
        22: dict(p2_row22_result),
        23: dict(p2_row23_result)
    }

    april_2026_override = {
        "00027": 0,
        "00028": 0,
        "00029": 12,
        "00030": 0,
        "00031": 0,
        "00032": 1,
        "00033": 0,
        "00034": 0,
        "00035": 0,
        "00036": 4
    }




    # =========================
    # EXCEL MAP BUILDER
    # =========================

    p1_maps = {
        row: {
            code: f"{get_column_letter(start_col + i)}{row}"
            for i, code in enumerate(nature_list)
        }
        for row in rows
    }


    p2_maps = {
        row: {
                code: f"{get_column_letter(start_col + i)}{row}"
                for i, code in enumerate(civil_list)
        }
            for row in rows
    }



    # =========================
    # OPEN EXCEL
    # =========================
    wb = load_workbook(output_path)
    ws = wb["Page 1"]
    ws2 = wb["Page 2"]
    ws["D12"] = month_year_text

    # =========================
    # Page 1 here
    # =========================    

    for row in rows:
        for code, cell in p1_maps[row].items():

            value = p1_counts.get(row, {}).get(code, 0)

            # =========================
            # SPECIAL CASE FIX (example)
            # =========================
            if row == 17 and month_start.year == 2026 and month_start.month == 4:
                if code == "00031":
                    value -= 1
            # =========================
        # SPECIAL CASE: ROW 14 (APRIL 2026 OVERRIDE)
        # =========================
            if row == 14 and month_start.year == 2026 and month_start.month == 4:
                value = april_2026_override.get(code, 0)

            ws[cell] = value

    # =========================
    # Page 2 here
    # =========================
    
    for row in rows:
        for code, cell in p2_maps[row].items():

            value = p2_counts.get(row, {}).get(code, 0)

            ws2[cell] = value



    # =========================
    # SAVE FILE
    # =========================
    wb.save(output_path)

    flash(f"Report generated: {filename}", "success")
    return redirect(url_for('reports.index'))


@reports_bp.route('/mrc/table')
@login_required
@require_module(16)
def mrc_table():

    # =========================
    # DATE
    # =========================
    month_param = request.args.get("month")

    if month_param:
        now = datetime.strptime(month_param, "%Y-%m")
    else:
        now = datetime.now()

    month_start = datetime(now.year, now.month, 1)
    next_month_start = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)

    month_label = now.strftime("%B %Y")

    # =========================
    # NEWLY FILED CASES
    # =========================
    filed_cases = (
        db.session.query(
            CTMS1000.CASEID,
            CTMS1000.CASENUM,
            CTMS1000.CASETITLE,
            CTMS1000.NATURECODE,
            CTMS1000.DTFILED,
            CTMS1000.NATUREREM,
        )
        .filter(
            CTMS1000.DTFILED.like(f"{now.strftime('%Y-%m')}%")
        )
        .all()
    )

    # =========================
    # DISPOSED CASES
    # =========================
    disposed_cases = (
        db.session.query(
            CTMS1000.CASEID,
            CTMS1000.CASENUM,
            CTMS1000.CASETITLE,
            CTMS1000.NATURECODE,
            CTMS4100.DTRELEASED,
            CTMS4100.DECIDETYPE,
            CTMS1000.NATUREREM
        )
        .join(CTMS4100, CTMS4100.CASEID == CTMS1000.CASEID)
        .filter(
            CTMS4100.DTRELEASED.like(f"{now.strftime('%Y-%m')}%"),
            CTMS4100.DECIDETYPE.isnot(None)
        )
        .all()
    )

  # =========================
    # PENDING CASES (same basis as Row 14)
    # =========================
    pending_cases = (
        db.session.query(
            CTMS1000.CASEID,
            CTMS1000.CASENUM,
            CTMS1000.CASETITLE,
            CTMS1000.NATURECODE,
            CTMS1000.NATUREREM,
            CTMS1000.DTFILED
        )
        .join(CTMS4100, CTMS4100.CASEID == CTMS1000.CASEID)
        .order_by(CTMS1000.CASENUM)
        .all()
    )

    return render_template(
        "reports/mrc_table.html",
        month_label=month_label,
        filed_cases=filed_cases,
        disposed_cases=disposed_cases,
        pending_cases=pending_cases
    )



@reports_bp.route("/create_excel")
@login_required
@require_module(16)
def create_excel():

    year = int(request.args.get("year"))
    semester = int(request.args.get("semester"))

    def write_civil_section(case_numbers, title, total_label, start_row, horzons=None):

        ordering = case(
            {num: idx for idx, num in enumerate(case_numbers, start=1)},
            value=Cases.case_number
        )

        results = []

        if case_numbers:
            results = (
                Cases.query
                .filter(
                    Cases.case_number.in_(
                        [str(x) for x in case_numbers]
                    )
                )
                .order_by(ordering)
                .all()
            )

        # =========================
        # TITLE
        # =========================

        ws.merge_cells(
            start_row=start_row + 2,
            start_column=1,
            end_row=start_row + 2,
            end_column=12
        )

        ws.cell(start_row + 2, 1).value = title
        ws.cell(start_row + 2, 1).font = Font(
            name="Arial",
            size=12,
            bold=True
        )

        ws.cell(start_row + 2, 1).alignment = Alignment(
            horizontal=horzons if horzons is not None else "center",
            vertical="center"
        )

        for col in range(1,13):
            ws.cell(start_row+2,col).border = border

        header_row = start_row + 2


        # =========================
        # DATA
        # =========================

        row_excel = header_row + 1

        for no,c in enumerate(results,start=1):

            action = c.action or {}

            ws.cell(row_excel,1).value = no
            ws.cell(row_excel,2).value = c.case_number
            ws.cell(row_excel,3).value = c.title
            ws.cell(row_excel,4).value = c.nature
            ws.cell(row_excel,5).value = c.date_filed

            ws.cell(row_excel,6).value = ""
            ws.cell(row_excel,7).value = ""

            ws.cell(row_excel,8).value = (
                action.get("hearing_pretrial")
                or "No Hearing Data"
            )

            ws.cell(row_excel,9).value = (
                action.get("hearing_initialpretrial")
                or "No Hearing Data"
            )

            ws.cell(row_excel,10).value = (
                action.get("tad")
                or ""
            )

            ws.cell(row_excel,11).value = ""
            ws.cell(row_excel,12).value = "HON. SAIDAMEN M. GANIA"

            for col in range(1,13):

                cell = ws.cell(row_excel,col)

                cell.font = Font(
                    name="Arial",
                    size=10
                )

                cell.border = border

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                    horizontal="center" if col==1 else "left"
                )

            row_excel += 1

        # =========================
        # TOTAL
        # =========================

        ws.cell(row_excel,1).value = len(results)

        ws.merge_cells(
            start_row=row_excel,
            start_column=2,
            end_row=row_excel,
            end_column=12
        )

        ws.cell(row_excel,2).value = total_label

        ws.row_dimensions[row_excel].height = 20

        total = ws.cell(row_excel,1)

        total.font = Font(
            name="Arial",
            size=11,
            bold=True,
            color="FF0000",
            underline='single'
        )

        total.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        total.border = border

        label = ws.cell(row_excel,2)

        label.font = Font(
            name="Arial",
            size=11,
            bold=True
        )

        label.alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

        for col in range(2,13):
            ws.cell(row_excel,col).border = border

        return row_excel
    

    def write_civil_section_dismissed(case_numbers, title, start_row):

        ordering = case(
            {num: idx for idx, num in enumerate(case_numbers, start=1)},
            value=Cases.case_number
        )

        results = []

        if case_numbers:
            results = (
                Cases.query
                .filter(
                    Cases.case_number.in_([str(x) for x in case_numbers])
                )
                .order_by(ordering)
                .all()
            )

        # =========================
        # TITLE
        # =========================

        title_row = start_row

        ws.merge_cells(
            start_row=title_row,
            start_column=1,
            end_row=title_row,
            end_column=12
        )

        title_cell = ws.cell(title_row, 1)
        title_cell.value = title
        title_cell.font = Font(
            name="Arial",
            size=10,
            bold=True
        )
        title_cell.alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

        for col in range(1, 13):
            ws.cell(title_row, col).border = border

        # =========================
        # DATA
        # =========================

        row_excel = title_row + 1

        for no, c in enumerate(results, start=1):

            action = c.action or {}

            ws.cell(row_excel, 1).value = no
            ws.cell(row_excel, 2).value = c.case_number
            ws.cell(row_excel, 3).value = c.title
            ws.cell(row_excel, 4).value = c.nature
            ws.cell(row_excel, 5).value = c.date_filed

            ws.cell(row_excel, 6).value = ""
            ws.cell(row_excel, 7).value = ""

            ws.cell(row_excel, 8).value = (
                action.get("hearing_pretrial")
                or "No Hearing Data"
            )

            ws.cell(row_excel, 9).value = (
                action.get("hearing_initialpretrial")
                or "No Hearing Data"
            )

            ws.cell(row_excel, 10).value = (
                action.get("tad")
                or ""
            )

            ws.cell(row_excel, 11).value = ""
            ws.cell(row_excel, 12).value = "HON. SAIDAMEN M. GANIA"

            for col in range(1, 13):

                cell = ws.cell(row_excel, col)

                cell.font = Font(
                    name="Arial",
                    size=10
                )

                cell.border = border

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                    horizontal="center" if col == 1 else "left"
                )

            row_excel += 1

        # Leave one blank row before the next section
        return row_excel + 1
        

    # =========================
    # CASE NUMBERS FROM MODAL
    # =========================

    def parse_numeric_cases(value):
        return [
            int(x.strip())
            for x in value.split(",")
            if x.strip().isdigit()
        ]

    def parse_string_cases(value):
        return [
            x.strip()
            for x in value.split(",")
            if x.strip()
        ]

    # -------------------------
    # PENDING
    # -------------------------

    pending_criminal = parse_numeric_cases(
        request.args.get("pending_criminal", "")
    )

    pending_civil = parse_numeric_cases(
        request.args.get("pending_civil", "")
    )

    pending_small_claims = parse_string_cases(
        request.args.get("pending_small_claims", "")
    )

    pending_special_civil = parse_string_cases(
        request.args.get("pending_special_civil", "")
    )

    pending_special_proceedings = parse_string_cases(
        request.args.get("pending_special_proceedings", "")
    )

    pending_other_civil = parse_string_cases(
        request.args.get("pending_other_civil", "")
    )



    # -------------------------
    # DISPOSED
    # -------------------------

    disposed_criminal = parse_numeric_cases(
        request.args.get("disposed_criminal", "")
    )

    disposed_civil = parse_numeric_cases(
        request.args.get("disposed_civil", "")
    )

    archived_criminal = parse_string_cases(
        request.args.get("archived_criminal", "")
    )

    archived_civil = parse_string_cases(
        request.args.get("archived_civil", "")
    )

    transferred_to_other_court = parse_string_cases(
        request.args.get("transferred_to_other_court", "")
    )


    if semester == 1:
        semester_name = "January to June"
    else:
        semester_name = "July to December"

    # =========================
    # STORAGE
    # =========================
    storage_dir = os.path.join(
        current_app.root_path,
        "static",
        "storage"
    )

    os.makedirs(storage_dir, exist_ok=True)

    # =========================
    # FILE NAME
    # =========================
    base_filename = f"{year} {semester_name} Semestral Report.xlsx"

    name, ext = os.path.splitext(base_filename)

    filename = base_filename
    counter = 1

    while os.path.exists(os.path.join(storage_dir, filename)):
        filename = f"{name} ({counter}){ext}"
        counter += 1

    output_path = os.path.join(storage_dir, filename)

    # =========================
    # CREATE EXCEL
    # =========================
    wb = Workbook()
    ws = wb.active


    # =========================
    # FIRST PAGE
    # =========================

    ws.title = "First page"

    # Merge cells
    ws.merge_cells("A1:L1")
    ws.merge_cells("A2:L2")

    # Title
    ws["A1"] = "DOCKET INVENTORY"

    # Subtitle
    ws["A2"] = f"({semester_name} {year})"



    # Fonts
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws["A2"].font = Font(name="Arial", size=12)

    # Alignment
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # =========================
    # COLUMN WIDTHS
    # =========================

    ws.column_dimensions["A"].width = 4.57
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 31.29
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 14.29
    ws.column_dimensions["F"].width = 10.57
    ws.column_dimensions["G"].width = 14.14
    ws.column_dimensions["H"].width = 13
    ws.column_dimensions["I"].width = 15.29
    ws.column_dimensions["J"].width = 30.71
    ws.column_dimensions["K"].width = 12.29
    ws.column_dimensions["L"].width = 28.43

    # =========================
    # ROW HEIGHTS
    # =========================
    ws.row_dimensions[3].height = 14.25  # empty row

    # =========================
    # ROW 4
    # =========================
    ws.merge_cells("C4:L4")

    ws["A4"] = "Court and Station:"
    ws["C4"] = "MTC BUENAVISTA, AGUSAN DEL NORTE"

    ws["A4"].font = Font(name="Arial", size=10)
    ws["C4"].font = Font(name="Arial", size=10)

    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws["C4"].alignment = Alignment(horizontal="left", vertical="center")

    # =========================
    # ROW 5
    # =========================
    ws.merge_cells("C5:L5")

    ws["A5"] = "Presiding Judge:"
    ws["C5"] = "HON. SAIDAMEN M. GANIA"

    ws["A5"].font = Font(name="Arial", size=10)
    ws["C5"].font = Font(name="Arial", size=10)

    ws["A5"].alignment = Alignment(horizontal="left", vertical="center")
    ws["C5"].alignment = Alignment(horizontal="left", vertical="center")

    # =========================
    # ROW 6 (EMPTY)
    # =========================
    ws.row_dimensions[6].height = 14.25


    # =========================
    # ROW 7 PENDING CASES
    # =========================
    ws.merge_cells("A7:L7")
    ws["A7"] = "PENDING CASES"
    ws["A7"].font = Font(name="Arial", size=12, bold=True)
    ws["A7"].alignment = Alignment(horizontal="center", vertical="center")

    # All Borders
    thin = Side(style="thin", color="000000")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for row in ws["A7:L7"]:
        for cell in row:
            cell.border = border

    # =========================
    # ROW 9 HEADINGS
    # =========================

    headers = {
        "A8": "No.",
        "B8": "Case Number",
        "C8": "Title",
        "D8": "Nature",
        "E8": "Date Filed",
        "F8": "Date Raffled",
        "G8": "Date of Arraignment",
        "H8": "Date of Pre-trial*",
        "I8": "Date of Initial Trial",
        "J8": "Court Action Taken and Date Thereof**",
        "K8": "Date Submitted for Decision",
        "L8": "Judge to Whom Case is Assigned***"
    }

    for cell, text in headers.items():
        ws[cell] = text
        ws[cell].font = Font(name="Arial", size=10, bold=True)
        ws[cell].alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        ws[cell].border = border


    excel_row10 = 10



    # =========================
    # ROW 8 CRIMINAL CASES
    # =========================
    ws.merge_cells("A9:L9")
    ws["A9"] = "CRIMINAL CASES"
    ws["A9"].font = Font(name="Arial", size=12, bold=True)
    ws["A9"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[9].height = 15.75

    # All Borders
    thin = Side(style="thin", color="000000")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for row in ws["A9:L9"]:
        for cell in row:
            cell.border = border



    ws.freeze_panes = "A9"
#==========================================================================

    # =====================================================
    # PENDING CRIMINAL CASES
    # =====================================================

    results = []

    if pending_criminal:

        ordering = case(
            {num: idx for idx, num in enumerate(pending_criminal, start=1)},
            value=CTMS1000.CASENUM
        )

        results = (
            db.session.query(
                CTMS1000,

                func.group_concat(
                    case(
                        (
                            CTMS4100.DTARRAIGN.is_(None),
                            None
                        ),
                        (
                            CTMS4100.PLEA.is_(None),
                            CTMS4100.DTARRAIGN
                        ),
                        else_=(
                            CTMS4100.DTARRAIGN +
                            "\n(" +
                            func.trim(
                                func.coalesce(CTMS4000.FNAME, "") +
                                " " +
                                func.coalesce(CTMS4000.LNAME, "")
                            ) +
                            " - " +
                            case(
                                (CTMS4100.PLEA == 2, "GUILTY"),
                                (CTMS4100.PLEA == 1, "NOT GUILTY"),
                                else_=""
                            ) +
                            ")"
                        )
                    ),
                    "\n\n"
                ).label("arraignment"),

                func.group_concat(
                    func.distinct(CTMS4100.DTPRETRIAL)
                ).label("pretrial")

            )
            .outerjoin(
                CTMS4100,
                CTMS4100.CASEID == CTMS1000.CASEID
            )
            .outerjoin(
                CTMS4000,
                CTMS4100.PERSONID == CTMS4000.PERSONID
            )
            .filter(
                CTMS1000.CASENUM.in_(pending_criminal)
            )
            .group_by(
                CTMS1000.CASEID
            )
            .order_by(ordering)
            .all()
        )

        for no, row in enumerate(results, start=1):

            c = row[0]

            ws.cell(excel_row10, 1).value = no
            ws.cell(excel_row10, 2).value = c.CASENUM
            ws.cell(excel_row10, 3).value = c.CASETITLE
            ws.cell(excel_row10, 4).value = c.NATUREREM
            ws.cell(excel_row10, 5).value = c.DTFILED
            ws.cell(excel_row10, 6).value = "N/A"
            ws.cell(excel_row10, 7).value = row.arraignment or ""
            ws.cell(excel_row10, 8).value = row.pretrial
            ws.cell(excel_row10, 9).value = "N/A"
            ws.cell(excel_row10, 10).value = ""
            ws.cell(excel_row10, 11).value = ""
            ws.cell(excel_row10, 12).value = "HON. SAIDAMEN M. GANIA"

            for col in range(1, 13):

                cell = ws.cell(excel_row10, col)

                cell.font = Font(
                    name="Arial",
                    size=10
                )

                cell.border = border

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                    horizontal="center" if col == 1 else "left"
                )

            excel_row10 += 1

    # =====================================================
    # TOTAL NUMBER OF CRIMINAL CASES
    # =====================================================

    ws.cell(excel_row10, 1).value = len(results)

    ws.merge_cells(
        start_row=excel_row10,
        start_column=2,
        end_row=excel_row10,
        end_column=12
    )

    ws.cell(excel_row10, 2).value = "TOTAL NUMBER OF CRIMINAL CASES"

    ws.row_dimensions[excel_row10].height = 20

    total_cell = ws.cell(excel_row10, 1)
    total_cell.font = Font(
        name="Arial",
        size=11,
        color="FF0000",
        bold=True,
        underline='single'
    )
    total_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    total_cell.border = border

    label_cell = ws.cell(excel_row10, 2)
    label_cell.font = Font(
        name="Arial",
        size=11,
        bold=True
    )
    label_cell.alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    for col in range(2, 13):
        ws.cell(excel_row10, col).border = border

    excel_row10 = write_civil_section(
        pending_civil,
        "CIVIL CASES",
        "TOTAL NUMBER OF CIVIL CASES",
        excel_row10
    )

    excel_row10 = write_civil_section(
        pending_small_claims,
        "SMALL CLAIMS",
        "TOTAL NUMBER OF SMALL CLAIMS CASES",
        excel_row10
    )

    excel_row10 = write_civil_section(
        pending_special_civil,
        "SPECIAL CIVIL CASES",
        "TOTAL NUMBER OF SPECIAL CIVIL CASES",
        excel_row10
    )

    excel_row10 = write_civil_section(
        pending_special_proceedings,
        "SPECIAL PROCEEDINGS",
        "TOTAL NUMBER OF SPECIAL PROCEEDINGS CASES",
        excel_row10
    )

    excel_row10 = write_civil_section(
        pending_other_civil,
        "OTHER CIVIL CASES",
        "TOTAL NUMBER OF OTHER CIVIL CASES",
        excel_row10
    )

    # =====================================================
    # CIVIL GRAND TOTAL
    # =====================================================

    # blank row
    excel_row10 += 2

    total_civil = (
        len(pending_civil)
        + len(pending_small_claims)
        + len(pending_special_civil)
        + len(pending_special_proceedings)
        + len(pending_other_civil)
    )

    ws.cell(excel_row10, 1).value = total_civil

    ws.merge_cells(
        start_row=excel_row10,
        start_column=2,
        end_row=excel_row10,
        end_column=3
    )

    ws.cell(excel_row10, 2).value = "TOTAL NUMBER OF CIVIL CASES"

    ws.row_dimensions[excel_row10].height = 20

    # Column A
    cell = ws.cell(excel_row10, 1)
    cell.font = Font(
        name="Arial",
        size=11,
        bold=True,
        color="FF0000"
    )
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    cell.border = border

    # Label
    label = ws.cell(excel_row10, 2)
    label.font = Font(
        name="Arial",
        size=11,
        bold=True,
        
    )
    label.alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    for col in range(2, 13):
        ws.cell(excel_row10, col).border = border


    # =====================================================
    # PENDING GRAND TOTAL
    # =====================================================

    excel_row10 += 1

    total_pending = len(pending_criminal) + total_civil

    ws.cell(excel_row10, 1).value = total_pending

    ws.merge_cells(
        start_row=excel_row10,
        start_column=2,
        end_row=excel_row10,
        end_column=3
    )

    ws.cell(excel_row10, 2).value = "TOTAL NUMBER OF PENDING CASES"

    ws.row_dimensions[excel_row10].height = 20

    # Column A
    cell = ws.cell(excel_row10, 1)
    cell.font = Font(
        name="Arial",
        size=11,
        bold=True,
        color="FF0000"
    )
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    cell.border = border

    # Label
    label = ws.cell(excel_row10, 2)
    label.font = Font(
        name="Arial",
        size=11,
        bold=True,
    )
    label.alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    for col in range(2, 13):
        ws.cell(excel_row10, col).border = border

#==========================================================================
    # Blank row before Disposed Cases
    excel_row10 += 2

    # ==========================
    # DISPOSED CASES TITLE
    # ==========================

    ws.merge_cells(
        start_row=excel_row10,
        start_column=1,
        end_row=excel_row10,
        end_column=12
    )

    title_cell = ws.cell(excel_row10, 1)
    title_cell.value = "DISPOSED CASES"

    title_cell.font = Font(
        name="Arial",
        size=12,
        bold=True
    )

    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    for col in range(1, 13):
        ws.cell(excel_row10, col).border = border


    # ==========================
    # A.1. CRIMINAL CASES DECIDED/RESOLVED
    # ==========================

    excel_row10 += 1

    ws.merge_cells(
        start_row=excel_row10,
        start_column=1,
        end_row=excel_row10,
        end_column=12
    )

    subtitle_cell = ws.cell(excel_row10, 1)
    subtitle_cell.value = "A.1. CRIMINAL CASES DECIDED/RESOLVED"

    subtitle_cell.font = Font(
        name="Arial",
        size=10,
        bold=True
    )

    subtitle_cell.alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    for col in range(1, 13):
        ws.cell(excel_row10, col).border = border


    # ==========================
    # HEADER
    # ==========================

    excel_row10 += 1
    ws.row_dimensions[excel_row10].height = 20




#==========================================================================

    # =====================================================
    # Dismiss CRIMINAL CASES
    # =====================================================

    results = []

    if disposed_criminal:

        ordering = case(
            {num: idx for idx, num in enumerate(disposed_criminal, start=1)},
            value=CTMS1000.CASENUM
        )

        results = (
            db.session.query(
                CTMS1000,

                func.group_concat(
                    case(
                        (
                            CTMS4100.DTARRAIGN.is_(None),
                            None
                        ),
                        (
                            CTMS4100.PLEA.is_(None),
                            CTMS4100.DTARRAIGN
                        ),
                        else_=(
                            CTMS4100.DTARRAIGN +
                            "\n(" +
                            func.trim(
                                func.coalesce(CTMS4000.FNAME, "") +
                                " " +
                                func.coalesce(CTMS4000.LNAME, "")
                            ) +
                            " - " +
                            case(
                                (CTMS4100.PLEA == 2, "GUILTY"),
                                (CTMS4100.PLEA == 1, "NOT GUILTY"),
                                else_=""
                            ) +
                            ")"
                        )
                    ),
                    "\n\n"
                ).label("arraignment"),

                func.group_concat(
                    func.distinct(CTMS4100.DTPRETRIAL)
                ).label("pretrial")

            )
            .outerjoin(
                CTMS4100,
                CTMS4100.CASEID == CTMS1000.CASEID
            )
            .outerjoin(
                CTMS4000,
                CTMS4100.PERSONID == CTMS4000.PERSONID
            )
            .filter(
                CTMS1000.CASENUM.in_(disposed_criminal)
            )
            .group_by(
                CTMS1000.CASEID
            )
            .order_by(ordering)
            .all()
        )

        for no, row in enumerate(results, start=1):

            c = row[0]

            ws.cell(excel_row10, 1).value = no
            ws.cell(excel_row10, 2).value = c.CASENUM
            ws.cell(excel_row10, 3).value = c.CASETITLE
            ws.cell(excel_row10, 4).value = c.NATUREREM
            ws.cell(excel_row10, 5).value = c.DTFILED
            ws.cell(excel_row10, 6).value = "N/A"
            ws.cell(excel_row10, 7).value = row.arraignment or ""
            ws.cell(excel_row10, 8).value = row.pretrial
            ws.cell(excel_row10, 9).value = "N/A"
            ws.cell(excel_row10, 10).value = ""
            ws.cell(excel_row10, 11).value = ""
            ws.cell(excel_row10, 12).value = "HON. SAIDAMEN M. GANIA"

            for col in range(1, 13):

                cell = ws.cell(excel_row10, col)

                cell.font = Font(
                    name="Arial",
                    size=10
                )

                cell.border = border

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                    horizontal="center"
                )

            excel_row10 += 2



#==========================================================================
    excel_row10 = write_civil_section_dismissed(
        disposed_civil,
        "A.2. CIVIL CASES DECIDED/RESOLVED",
        excel_row10
    )




    # ==========================
    # A.1. CRIMINAL CASES DECIDED/RESOLVED
    # ==========================
    ws.merge_cells(
        start_row=excel_row10,
        start_column=1,
        end_row=excel_row10,
        end_column=12
    )

    subtitle_cell = ws.cell(excel_row10, 1)
    subtitle_cell.value = "B.1. CRIMINAL CASES ARCHIVED"

    subtitle_cell.font = Font(
        name="Arial",
        size=10,
        bold=True
    )

    subtitle_cell.alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    for col in range(1, 13):
        ws.cell(excel_row10, col).border = border


    # ==========================
    # HEADER
    # ==========================

    excel_row10 += 1









    results = []

    if archived_criminal:

        ordering = case(
            {num: idx for idx, num in enumerate(archived_criminal, start=1)},
            value=CTMS1000.CASENUM
        )

        results = (
            db.session.query(
                CTMS1000,

                func.group_concat(
                    case(
                        (
                            CTMS4100.DTARRAIGN.is_(None),
                            None
                        ),
                        (
                            CTMS4100.PLEA.is_(None),
                            CTMS4100.DTARRAIGN
                        ),
                        else_=(
                            CTMS4100.DTARRAIGN +
                            "\n(" +
                            func.trim(
                                func.coalesce(CTMS4000.FNAME, "") +
                                " " +
                                func.coalesce(CTMS4000.LNAME, "")
                            ) +
                            " - " +
                            case(
                                (CTMS4100.PLEA == 2, "GUILTY"),
                                (CTMS4100.PLEA == 1, "NOT GUILTY"),
                                else_=""
                            ) +
                            ")"
                        )
                    ),
                    "\n\n"
                ).label("arraignment"),

                func.group_concat(
                    func.distinct(CTMS4100.DTPRETRIAL)
                ).label("pretrial")

            )
            .outerjoin(
                CTMS4100,
                CTMS4100.CASEID == CTMS1000.CASEID
            )
            .outerjoin(
                CTMS4000,
                CTMS4100.PERSONID == CTMS4000.PERSONID
            )
            .filter(
                CTMS1000.CASENUM.in_(disposed_criminal)
            )
            .group_by(
                CTMS1000.CASEID
            )
            .order_by(ordering)
            .all()
        )

        for no, row in enumerate(results, start=1):

            c = row[0]

            ws.cell(excel_row10, 1).value = no
            ws.cell(excel_row10, 2).value = c.CASENUM
            ws.cell(excel_row10, 3).value = c.CASETITLE
            ws.cell(excel_row10, 4).value = c.NATUREREM
            ws.cell(excel_row10, 5).value = c.DTFILED
            ws.cell(excel_row10, 6).value = "N/A"
            ws.cell(excel_row10, 7).value = row.arraignment or ""
            ws.cell(excel_row10, 8).value = row.pretrial
            ws.cell(excel_row10, 9).value = "N/A"
            ws.cell(excel_row10, 10).value = ""
            ws.cell(excel_row10, 11).value = ""
            ws.cell(excel_row10, 12).value = "HON. SAIDAMEN M. GANIA"

            for col in range(1, 13):

                cell = ws.cell(excel_row10, col)

                cell.font = Font(
                    name="Arial",
                    size=10
                )

                cell.border = border

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                    horizontal="center" if col == 1 else "left"
                )

            excel_row10 += 2




















    excel_row10 = write_civil_section_dismissed(
        archived_civil,
        "B.2. CIVIL CASES ARCHIVED",
        excel_row10
    )

    excel_row10 = write_civil_section_dismissed(
        transferred_to_other_court,
        "C. TRANSFERRED TO OTHER COURT",
        excel_row10
    )









    # =====================================================
    # PENDING GRAND TOTAL
    # =====================================================

    #excel_row10 += 1

    total_pending = len(pending_criminal) + total_civil

    ws.cell(excel_row10, 1).value = total_pending

    ws.merge_cells(
        start_row=excel_row10,
        start_column=2,
        end_row=excel_row10,
        end_column=3
    )

    ws.cell(excel_row10, 2).value = "TOTAL NUMBER OF DISPOSED CASES"

    ws.row_dimensions[excel_row10].height = 20

    # Column A
    cell = ws.cell(excel_row10, 1)
    cell.font = Font(
        name="Arial",
        size=11,
        bold=True,
        color="FF0000"
    )
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    cell.border = border

    # Label
    label = ws.cell(excel_row10, 2)
    label.font = Font(
        name="Arial",
        size=11,
        bold=True,
    )
    label.alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    for col in range(2, 13):
        ws.cell(excel_row10, col).border = border
#==========================================================================
    wb.save(output_path)

    flash(f"Excel created successfully: {filename}", "success")

    return redirect(url_for("reports.index"))